"""Generate project/Data/TestData/E2E.xlsx — EligibilityDecision sheet."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "EligibilityDecision"

headers = [
    "TestCaseId",
    "Description",
    "CaseCategory",
    "ComplianceAssessment",
    "Attribute_A",
    "Attribute_B",
    "Attribute_C",
    "Indicator_X",
    "Detail_X",
    "SecondaryIndicator",
    "Expected_DecisionValue",
    "RuleGroup",
    "Notes",
]

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="264E8C")
pos_fill    = PatternFill("solid", fgColor="C6EFCE")
neg_fill    = PatternFill("solid", fgColor="FFDDC1")
open_fill   = PatternFill("solid", fgColor="FFF2CC")
thin        = Side(style="thin")
border      = Border(left=thin, right=thin, top=thin, bottom=thin)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = border

NA = ""

# fmt: off
rows = [
    # ── Rule Group 1: CaseCategory = Event_A or Event_B ─────────────────────
    # Positive paths
    ("RG1-P-001", "Compliant -> Positive",
     "Event_A", "Compliant", NA, NA, NA, NA, NA, False,
     "Positive", 1, "Happy path"),
    ("RG1-P-002", "Event_B + Compliant -> Positive",
     "Event_B", "Compliant", NA, NA, NA, NA, NA, False,
     "Positive", 1, ""),
    ("RG1-P-003", "Non_Compliant + Verif_NA + Attr_B matches -> Positive",
     "Event_A", "Non_Compliant", "Verification_not_available", "No_or_low_impact", NA, NA, NA, False,
     "Positive", 1, "Attr_C not applicable"),
    ("RG1-P-004", "Non_Compliant + Verif_NA + Attr_C matches -> Positive",
     "Event_A", "Non_Compliant", "Verification_not_available", NA, "No_or_low_impact", NA, NA, False,
     "Positive", 1, "Attr_B not applicable"),
    ("RG1-P-005", "Non_Compliant + Verif_NA + both Attr match -> Positive",
     "Event_B", "Non_Compliant", "Verification_not_available", "No_or_low_impact", "No_or_low_impact", NA, NA, False,
     "Positive", 1, "Both Attr_B and Attr_C = No_or_low_impact"),
    ("RG1-P-006", "Non_Compliant + Special_case + Attr_B matches -> Positive",
     "Event_A", "Non_Compliant", "Special_case_without_certified_component", "No_or_low_impact", NA, NA, NA, False,
     "Positive", 1, ""),
    ("RG1-P-007", "Non_Compliant + Special_case + Attr_C matches -> Positive",
     "Event_B", "Non_Compliant", "Special_case_without_certified_component", NA, "No_or_low_impact", NA, NA, False,
     "Positive", 1, ""),
    ("RG1-P-008", "ComplianceAssessment empty -> Positive",
     "Event_A", NA, NA, NA, NA, NA, NA, False,
     "Positive", 1, "Empty ComplianceAssessment sentinel"),
    ("RG1-P-009", "Event_B + ComplianceAssessment empty -> Positive",
     "Event_B", NA, NA, NA, NA, NA, NA, False,
     "Positive", 1, "Empty ComplianceAssessment sentinel"),
    ("RG1-P-010", "SecondaryIndicator override -> Positive",
     "Event_A", "Non_Compliant", "OtherAttribute", "OtherImpact", "OtherImpact", NA, NA, True,
     "Positive", 1, "SecondaryIndicator=True overrides all other conditions"),
    ("RG1-P-011", "Event_B + SecondaryIndicator override -> Positive",
     "Event_B", "Non_Compliant", "OtherAttribute", NA, NA, NA, NA, True,
     "Positive", 1, "SecondaryIndicator=True overrides all other conditions"),
    # Negative paths
    ("RG1-N-001", "Non_Compliant + wrong Attr_A -> Negative",
     "Event_A", "Non_Compliant", "OtherAttribute", "No_or_low_impact", "No_or_low_impact", NA, NA, False,
     "Negative", 1, "Attr_A value not in allowed set"),
    ("RG1-N-002", "Non_Compliant + Verif_NA + no Attr match -> Negative",
     "Event_A", "Non_Compliant", "Verification_not_available", "OtherImpact", "OtherImpact", NA, NA, False,
     "Negative", 1, "Neither Attr_B nor Attr_C = No_or_low_impact"),
    ("RG1-N-003", "Non_Compliant + Special_case + no Attr match -> Negative",
     "Event_B", "Non_Compliant", "Special_case_without_certified_component", "OtherImpact", "OtherImpact", NA, NA, False,
     "Negative", 1, ""),
    # Open-question variants (AND vs OR for Attr_B / Attr_C)
    ("RG1-OQ-001", "[OPEN] Verif_NA + only Attr_B: AND interp -> Negative?",
     "Event_A", "Non_Compliant", "Verification_not_available", "No_or_low_impact", "OtherImpact", NA, NA, False,
     "?", 1, "Result depends on AND/OR resolution for Attr_B/Attr_C"),
    ("RG1-OQ-002", "[OPEN] Verif_NA + only Attr_C: AND interp -> Negative?",
     "Event_A", "Non_Compliant", "Verification_not_available", "OtherImpact", "No_or_low_impact", NA, NA, False,
     "?", 1, "Result depends on AND/OR resolution for Attr_B/Attr_C"),

    # ── Rule Group 2: CaseCategory = Event_C or Event_D ─────────────────────
    # Positive paths
    ("RG2-P-001", "Indicator_X=Y + Noticeable_irregularity -> Positive",
     "Event_C", NA, NA, NA, NA, "Y", "Noticeable_irregularity", False,
     "Positive", 2, ""),
    ("RG2-P-002", "Indicator_X=Y + Detail_X empty -> Positive",
     "Event_C", NA, NA, NA, NA, "Y", NA, False,
     "Positive", 2, "Empty Detail_X sentinel"),
    ("RG2-P-003", "Event_D + Indicator_X=Y + Noticeable_irregularity -> Positive",
     "Event_D", NA, NA, NA, NA, "Y", "Noticeable_irregularity", False,
     "Positive", 2, ""),
    ("RG2-P-004", "Event_D + Indicator_X=Y + Detail_X empty -> Positive",
     "Event_D", NA, NA, NA, NA, "Y", NA, False,
     "Positive", 2, ""),
    ("RG2-P-005", "Indicator_X=N + Compliant -> Positive",
     "Event_C", "Compliant", NA, NA, NA, "N", NA, False,
     "Positive", 2, ""),
    ("RG2-P-006", "Event_D + Indicator_X=N + Compliant -> Positive",
     "Event_D", "Compliant", NA, NA, NA, "N", NA, False,
     "Positive", 2, ""),
    ("RG2-P-007", "Indicator_X=N + Non_Compliant + Verif_NA + Attr_B -> Positive",
     "Event_C", "Non_Compliant", "Verification_not_available", "No_or_low_impact", NA, "N", NA, False,
     "Positive", 2, ""),
    ("RG2-P-008", "Indicator_X=N + Non_Compliant + Verif_NA + Attr_C -> Positive",
     "Event_D", "Non_Compliant", "Verification_not_available", NA, "No_or_low_impact", "N", NA, False,
     "Positive", 2, ""),
    ("RG2-P-009", "Indicator_X=N + Non_Compliant + Special_case + Attr_B -> Positive",
     "Event_C", "Non_Compliant", "Special_case_without_certified_component", "No_or_low_impact", NA, "N", NA, False,
     "Positive", 2, ""),
    ("RG2-P-010", "Indicator_X=N + Non_Compliant + Special_case + Attr_C -> Positive",
     "Event_D", "Non_Compliant", "Special_case_without_certified_component", NA, "No_or_low_impact", "N", NA, False,
     "Positive", 2, ""),
    ("RG2-P-011", "Indicator_X=N + ComplianceAssessment empty -> Positive",
     "Event_C", NA, NA, NA, NA, "N", NA, False,
     "Positive", 2, "Empty ComplianceAssessment sentinel"),
    ("RG2-P-012", "Event_D + Indicator_X=N + ComplianceAssessment empty -> Positive",
     "Event_D", NA, NA, NA, NA, "N", NA, False,
     "Positive", 2, ""),
    ("RG2-P-013", "SecondaryIndicator override -> Positive",
     "Event_C", "Non_Compliant", "OtherAttribute", "OtherImpact", "OtherImpact", "N", NA, True,
     "Positive", 2, "SecondaryIndicator=True overrides all other conditions"),
    ("RG2-P-014", "Event_D + SecondaryIndicator override -> Positive",
     "Event_D", "Non_Compliant", "OtherAttribute", NA, NA, "N", NA, True,
     "Positive", 2, ""),
    # Negative paths
    ("RG2-N-001", "Indicator_X=N + Non_Compliant + wrong Attr_A -> Negative",
     "Event_C", "Non_Compliant", "OtherAttribute", "No_or_low_impact", "No_or_low_impact", "N", NA, False,
     "Negative", 2, "Attr_A value not in allowed set"),
    ("RG2-N-002", "Indicator_X=N + Non_Compliant + Verif_NA + no Attr match -> Negative",
     "Event_D", "Non_Compliant", "Verification_not_available", "OtherImpact", "OtherImpact", "N", NA, False,
     "Negative", 2, ""),
    ("RG2-N-003", "Indicator_X=N + Non_Compliant + Special_case + no Attr match -> Negative",
     "Event_C", "Non_Compliant", "Special_case_without_certified_component", "OtherImpact", "OtherImpact", "N", NA, False,
     "Negative", 2, ""),
    # Open-question variants RG2
    ("RG2-OQ-001", "[OPEN] Indicator_X=N + Verif_NA + only Attr_B: AND interp -> Negative?",
     "Event_C", "Non_Compliant", "Verification_not_available", "No_or_low_impact", "OtherImpact", "N", NA, False,
     "?", 2, "Result depends on AND/OR resolution for Attr_B/Attr_C"),
    ("RG2-OQ-002", "[OPEN] Indicator_X=N + Verif_NA + only Attr_C: AND interp -> Negative?",
     "Event_C", "Non_Compliant", "Verification_not_available", "OtherImpact", "No_or_low_impact", "N", NA, False,
     "?", 2, "Result depends on AND/OR resolution for Attr_B/Attr_C"),
]
# fmt: on

expected_col = headers.index("Expected_DecisionValue") + 1

for r_idx, row in enumerate(rows, 2):
    for c_idx, val in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.border = border
        cell.alignment = Alignment(wrap_text=True)
    expected = row[10]
    result_cell = ws.cell(row=r_idx, column=expected_col)
    if expected == "Positive":
        result_cell.fill = pos_fill
    elif expected == "Negative":
        result_cell.fill = neg_fill
    elif expected == "?":
        result_cell.fill = open_fill

col_widths = [14, 52, 12, 22, 38, 16, 16, 12, 24, 18, 24, 10, 52]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "C2"
ws.auto_filter.ref = ws.dimensions

out = Path("project/Data/TestData/E2E.xlsx")
out.parent.mkdir(parents=True, exist_ok=True)
wb.save(out)
print(f"Written {len(rows)} rows -> {out}")
